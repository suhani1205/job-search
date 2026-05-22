import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Job Tracker"

headers = [
    "Job Title",
    "Company Name",
    "Location",
    "Job URL",
    "Date Posted",
    "Salary Range",
    "Key Requirements",
    "Match Score (1-10)",
    "Application Status",
    "Notes"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

jobs = [
    {
        "title": "Senior Software Engineer - Java, Spring Boot, Kafka",
        "company": "Wells Fargo",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.wellsfargojobs.com/en/jobs/r-542852/senior-software-engineer-java-spring-boot-kafka/",
        "date": "May 2026",
        "salary": "₹25-45 LPA (estimated)",
        "requirements": "Java, Spring Boot, Kafka, React.js, Microservices, REST APIs, Git",
        "score": 9,
        "status": "Not Applied",
        "notes": "Posting ends May 25, 2026. Strong match with Java/Spring Boot/Kafka stack."
    },
    {
        "title": "Senior Software Engineer - Java, Spring Boot, Microservices",
        "company": "Wells Fargo",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.wellsfargojobs.com/en/jobs/r-537664/senior-software-engineer-java-spring-boot-microservices/",
        "date": "May 2026",
        "salary": "₹25-45 LPA (estimated)",
        "requirements": "Java, Spring Boot, Microservices, REST APIs, Docker, Kubernetes",
        "score": 9,
        "status": "Not Applied",
        "notes": "4+ years Java Full Stack required. Excellent profile match."
    },
    {
        "title": "Java Engineer 3 (AWS)",
        "company": "Bread Financial India",
        "location": "Hybrid - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/view/java-engineer-3-aws-at-bread-financial-india-4401321331",
        "date": "May 2026",
        "salary": "₹20-35 LPA (estimated)",
        "requirements": "Java, AWS, Spring Boot, Microservices, SOLID Principles, Design Patterns",
        "score": 9,
        "status": "Not Applied",
        "notes": "AWS-focused role at fintech company. Strong skill overlap."
    },
    {
        "title": "Software Senior Engineer (Java, SpringBoot)",
        "company": "Barracuda Networks",
        "location": "Hybrid - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/view/software-senior-engineer-java-springboot-at-barracuda-4394621157",
        "date": "Apr 2026",
        "salary": "₹22-38 LPA (estimated)",
        "requirements": "Java, Spring Boot, Cloud, Data Protection, Cybersecurity",
        "score": 8,
        "status": "Not Applied",
        "notes": "Cloud-to-Cloud Backup team. Cybersecurity domain. Posted Apr 13, 2026."
    },
    {
        "title": "Senior Software Engineer - Java Springboot, PostgreSQL",
        "company": "UnitedHealth Group",
        "location": "Onsite - Chennai, Tamil Nadu",
        "url": "https://careers.unitedhealthgroup.com/job/chennai/senior-software-engineer-java-springboot-postgresql/34088/94804238272",
        "date": "May 7, 2026",
        "salary": "₹22-40 LPA (estimated)",
        "requirements": "Java, Spring Boot, PostgreSQL, RESTful APIs, Microservices",
        "score": 8,
        "status": "Not Applied",
        "notes": "Healthcare domain. RESTful APIs and microservices focus."
    },
    {
        "title": "Senior Software Engineer (Java Full Stack)",
        "company": "Caterpillar",
        "location": "Onsite - Bengaluru / Chennai",
        "url": "https://careers.caterpillar.com/en/jobs/r0000356686/senior-software-engineer-java-full-stack/",
        "date": "Apr 2026",
        "salary": "₹20-35 LPA (estimated)",
        "requirements": "Java/J2EE, Spring Boot/Spring MVC, Microservices, Multi-threaded Systems",
        "score": 8,
        "status": "Not Applied",
        "notes": "Manufacturing domain. Distributed high-availability web apps."
    },
    {
        "title": "Software Engineer - Java Backend",
        "company": "Caterpillar",
        "location": "Onsite - Bengaluru / Chennai",
        "url": "https://careers.caterpillar.com/en/jobs/r0000351358/software-engineer-java-backend/",
        "date": "Apr 2026",
        "salary": "₹15-28 LPA (estimated)",
        "requirements": "Java, Backend Development, REST APIs, Spring Boot",
        "score": 8,
        "status": "Not Applied",
        "notes": "Backend-focused role. Good for 3+ years experience level."
    },
    {
        "title": "Senior Software Engineer - Java Springboot",
        "company": "Rockwell Automation",
        "location": "Hybrid - Pune, Maharashtra",
        "url": "https://rockwellautomation.wd1.myworkdayjobs.com/en-US/External_Rockwell_Automation/job/Pune-India/Senior-Software-Engineer---Java-Springboot_R25-7962/apply",
        "date": "May 2026",
        "salary": "₹20-35 LPA (estimated)",
        "requirements": "Java, Spring Boot, Kubernetes, Jenkins, Maven, Git, SQL Server, Oracle",
        "score": 7,
        "status": "Not Applied",
        "notes": "MES/Pharma suite. Hybrid - Mon/Tue/Thu in Hinjewadi office."
    },
    {
        "title": "Senior Software Engineer - Java, Spring, Microservices, React/Angular",
        "company": "Wells Fargo",
        "location": "Onsite - Hyderabad, Telangana",
        "url": "https://www.wellsfargojobs.com/en/jobs/r-529306/senior-software-engineer-java-spring-microservices-react-angular/",
        "date": "May 2026",
        "salary": "₹25-45 LPA (estimated)",
        "requirements": "Java, Spring, Microservices, React/Angular, Kafka, Redis",
        "score": 9,
        "status": "Not Applied",
        "notes": "Full stack with microservices. Event-driven with Kafka."
    },
    {
        "title": "Senior Software Engineer - Full Stack Java, Spring Boot, ReactJS, AI",
        "company": "Wells Fargo",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.wellsfargojobs.com/en/jobs/r-538240/senior-software-engineer-full-stack-java-spring-boot-reactjs-ai/",
        "date": "May 2026",
        "salary": "₹28-48 LPA (estimated)",
        "requirements": "Java, Spring Boot, ReactJS, AI/ML Integration, Microservices",
        "score": 8,
        "status": "Not Applied",
        "notes": "AI-integrated role. React experience is a plus for this candidate."
    },
    {
        "title": "Software Engineer (Java, Kafka, Microservices)",
        "company": "Visa",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.instahyre.com/job-283679-software-engineer-at-visa-2-bangalore/",
        "date": "May 2026",
        "salary": "₹20-38 LPA (estimated)",
        "requirements": "Java, Python, Kafka, Microservices, REST, OpenTelemetry, Docker",
        "score": 8,
        "status": "Not Applied",
        "notes": "Payment Services & Transaction Platforms. Kafka streaming focus."
    },
    {
        "title": "Java Backend Developer",
        "company": "JPMorgan Chase & Co.",
        "location": "Hybrid - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/view/java-backend-developer-at-jpmorgan-chase-co-3200935938",
        "date": "May 2026",
        "salary": "₹25-50 LPA (estimated)",
        "requirements": "Java, Spring Boot, Microservices, AWS, REST APIs, SQL",
        "score": 9,
        "status": "Not Applied",
        "notes": "Top-tier financial institution. RSU + bonus packages."
    },
    {
        "title": "Senior Software Engineer II - Java",
        "company": "UnitedHealth Group",
        "location": "Onsite - Chennai, Tamil Nadu",
        "url": "https://careers.unitedhealthgroup.com/job/chennai/senior-software-engineer-ii-java/34088/93130764784",
        "date": "May 2026",
        "salary": "₹25-42 LPA (estimated)",
        "requirements": "Java, Spring Boot, Microservices, Cloud, CI/CD",
        "score": 8,
        "status": "Not Applied",
        "notes": "Healthcare tech. Senior-level position with growth potential."
    },
    {
        "title": "SDE 1/2 - Backend (Java Spring Boot)",
        "company": "CloudSEK",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.instahyre.com/job-281526-sde-1-2-backend-at-cloudsek-bangalore/",
        "date": "May 2026",
        "salary": "₹12-25 LPA (estimated)",
        "requirements": "Java/NodeJS/Go, Backend Development, REST APIs, Databases",
        "score": 7,
        "status": "Not Applied",
        "notes": "Cybersecurity startup. SDE-1/2 level matches experience."
    },
    {
        "title": "Sr. Software Engineer - Java/SpringBoot (MES)",
        "company": "Rockwell Automation (Kalypso)",
        "location": "Hybrid - Pune, Maharashtra",
        "url": "https://in.linkedin.com/jobs/view/sr-software-engineer-java-springboot-mes-at-rockwell-automation-4358319097",
        "date": "Apr 2026",
        "salary": "₹22-38 LPA (estimated)",
        "requirements": "Java, Spring Boot, ActiveMQ, Kubernetes, Jenkins, Maven, Oracle",
        "score": 7,
        "status": "Not Applied",
        "notes": "Manufacturing Execution Systems. Hybrid work model."
    },
    {
        "title": "Senior Software Engineering Lead - Java Backend",
        "company": "UnitedHealth Group",
        "location": "Onsite - Chennai, Tamil Nadu",
        "url": "https://careers.unitedhealthgroup.com/job/chennai/senior-software-engineering-lead-java-backend/34088/94041478624",
        "date": "May 2026",
        "salary": "₹30-50 LPA (estimated)",
        "requirements": "Java, Spring Boot, Microservices, Leadership, System Design",
        "score": 7,
        "status": "Not Applied",
        "notes": "Lead role - stretch opportunity. Healthcare domain."
    },
    {
        "title": "Software Engineer - Java",
        "company": "Caterpillar",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://careers.caterpillar.com/en/jobs/r0000354690/software-engineer-java/",
        "date": "May 2026",
        "salary": "₹15-28 LPA (estimated)",
        "requirements": "Java, Spring Boot, REST APIs, SQL, Git",
        "score": 8,
        "status": "Not Applied",
        "notes": "Good match for 3+ years experience. Enterprise environment."
    },
    {
        "title": "Senior Full Stack Engineer",
        "company": "UnitedHealth Group",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://careers.unitedhealthgroup.com/job/bengaluru/senior-full-stack-engineer/34088/94985594368",
        "date": "May 2026",
        "salary": "₹25-42 LPA (estimated)",
        "requirements": "Java, Spring Boot, React/Angular, Microservices, Cloud",
        "score": 8,
        "status": "Not Applied",
        "notes": "Full stack with Java backend focus. Bengaluru location."
    },
    {
        "title": "Backend Developer - Java, Spring Boot, AWS",
        "company": "Cutshort (Multiple Startups)",
        "location": "Remote - India",
        "url": "https://cutshort.io/jobs/backend-developer-jobs",
        "date": "May 2026",
        "salary": "₹15-35 LPA",
        "requirements": "Java 17+, Spring Boot, AWS (EC2, S3, IAM), Docker, Microservices",
        "score": 9,
        "status": "Not Applied",
        "notes": "Aggregated startup listings. Remote-friendly. Multiple openings."
    },
    {
        "title": "Backend Development Engineer - Java, Microservices",
        "company": "Instahyre (Multiple Companies)",
        "location": "Remote / Hybrid - India",
        "url": "https://www.instahyre.com/backend-development-jobs-in-anywhere-in-india/",
        "date": "May 2026",
        "salary": "₹18-40 LPA",
        "requirements": "Java, Spring Boot, Microservices, AWS, Docker, Kubernetes, REST APIs",
        "score": 9,
        "status": "Not Applied",
        "notes": "3986+ openings across India. Direct hiring manager contact."
    },
]

for row_idx, job in enumerate(jobs, 2):
    ws.cell(row=row_idx, column=1, value=job["title"])
    ws.cell(row=row_idx, column=2, value=job["company"])
    ws.cell(row=row_idx, column=3, value=job["location"])
    url_cell = ws.cell(row=row_idx, column=4, value=job["url"])
    url_cell.font = Font(color="0563C1", underline="single")
    ws.cell(row=row_idx, column=5, value=job["date"])
    ws.cell(row=row_idx, column=6, value=job["salary"])
    ws.cell(row=row_idx, column=7, value=job["requirements"])
    score_cell = ws.cell(row=row_idx, column=8, value=job["score"])
    score_cell.alignment = Alignment(horizontal="center")
    if job["score"] >= 9:
        score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif job["score"] >= 8:
        score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ws.cell(row=row_idx, column=9, value=job["status"])
    ws.cell(row=row_idx, column=10, value=job["notes"])
    
    for col_idx in range(1, 11):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx != 4:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

col_widths = {
    1: 45,  # Job Title
    2: 28,  # Company
    3: 30,  # Location
    4: 40,  # URL
    5: 15,  # Date Posted
    6: 22,  # Salary
    7: 55,  # Requirements
    8: 14,  # Match Score
    9: 16,  # Status
    10: 55, # Notes
}

for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.auto_filter.ref = f"A1:J{len(jobs) + 1}"
ws.freeze_panes = "A2"

now = datetime.utcnow()
filename = f"job_tracker_{now.strftime('%Y-%m-%d_%H')}.xlsx"
filepath = f"/home/user/job-search/{filename}"
wb.save(filepath)
print(f"Created: {filepath}")
print(f"Total jobs: {len(jobs)}")
