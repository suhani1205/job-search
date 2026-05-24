import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

now = datetime.now(timezone.utc)
filename = f"job_tracker_{now.strftime('%Y-%m-%d_%H')}.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Job Tracker"

headers = [
    "Job Title",
    "Company Name",
    "Location",
    "Job URL",
    "Date Posted",
    "Salary Range",
    "Key Requirements",
    "Match Score (1-10)",
    "Application Status",
    "Notes"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

jobs = [
    {
        "title": "SDE-2 (Backend, Java)",
        "company": "Flipkart",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.uplers.com/company/flipkart-5395",
        "date": "May 2026",
        "salary": "₹25,00,000 - ₹42,00,000",
        "requirements": "Java, Spring Boot, Microservices, AWS, Kafka, System Design, DSA",
        "score": 9,
        "status": "Not Applied",
        "notes": "Top product company; strong match with microservices & Kafka experience"
    },
    {
        "title": "Backend Engineer (Java/Spring Boot)",
        "company": "Razorpay",
        "location": "Hybrid - Bengaluru, Karnataka",
        "url": "https://razorpay.com/jobs/",
        "date": "May 2026",
        "salary": "₹22,00,000 - ₹38,00,000",
        "requirements": "Java, Spring Boot, Microservices, gRPC, AWS, MySQL, Docker",
        "score": 9,
        "status": "Not Applied",
        "notes": "Fintech; payment microservices align well with candidate's AWS + Spring Boot skills"
    },
    {
        "title": "Software Development Engineer II",
        "company": "Hindustan Times Digital",
        "location": "Hybrid - New Delhi",
        "url": "https://in.linkedin.com/jobs/spring-boot-jobs",
        "date": "May 2026",
        "salary": "₹18,00,000 - ₹30,00,000",
        "requirements": "Java, Spring Boot, REST APIs, AWS, Docker, CI/CD, Agile",
        "score": 8,
        "status": "Not Applied",
        "notes": "Media-tech; SDE-II level matches 3+ yrs experience"
    },
    {
        "title": "Java Microservices Developer",
        "company": "NTT DATA",
        "location": "Hybrid - Hyderabad, Telangana",
        "url": "https://www.naukri.com/java-microservice-spring-boot-jobs",
        "date": "May 2026",
        "salary": "₹12,00,000 - ₹22,00,000",
        "requirements": "Java 8+, Spring Boot, Microservices, AWS, Kafka, REST APIs, JUnit",
        "score": 9,
        "status": "Not Applied",
        "notes": "Global IT services; strong microservices focus matches profile perfectly"
    },
    {
        "title": "Senior Software Engineer (Java Backend)",
        "company": "Swiggy",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/java-spring-boot-developer-jobs-bengaluru",
        "date": "May 2026",
        "salary": "₹28,00,000 - ₹45,00,000",
        "requirements": "Java, Spring Boot, Kafka, AWS, Microservices, System Design, DynamoDB",
        "score": 9,
        "status": "Not Applied",
        "notes": "High-scale platform; DynamoDB + Kafka experience is a strong match"
    },
    {
        "title": "Backend Engineer (Java/Spring Boot)",
        "company": "CRED",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://wellfound.com/role/l/spring-boot/bangalore",
        "date": "May 2026",
        "salary": "₹25,00,000 - ₹50,00,000",
        "requirements": "Java, Spring Boot, Microservices, AWS, Docker, Kubernetes, REST APIs",
        "score": 8,
        "status": "Not Applied",
        "notes": "Fintech startup; premium compensation, needs strong system design skills"
    },
    {
        "title": "Java Developer (Spring Boot + AWS)",
        "company": "Tata Consultancy Services (TCS)",
        "location": "Hybrid - PAN India",
        "url": "https://www.naukri.com/spring-boot-aws-jobs",
        "date": "May 2026",
        "salary": "₹8,00,000 - ₹16,00,000",
        "requirements": "Java, Spring Boot, AWS (EC2, S3, Lambda), Microservices, MySQL, CI/CD",
        "score": 8,
        "status": "Not Applied",
        "notes": "Large-scale enterprise projects; multiple location options across India"
    },
    {
        "title": "SDE-2 (Cloud & Backend)",
        "company": "Meesho",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://www.productbased.in/",
        "date": "May 2026",
        "salary": "₹24,00,000 - ₹40,00,000",
        "requirements": "Java, Spring Boot, AWS, Kafka, Docker, Microservices, System Design",
        "score": 9,
        "status": "Not Applied",
        "notes": "Fast-growing e-commerce; high-traffic systems, great for cloud-native skills"
    },
    {
        "title": "Software Engineer - Backend (Java)",
        "company": "PhonePe",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/java-developer-spring-boot-jobs",
        "date": "May 2026",
        "salary": "₹22,00,000 - ₹38,00,000",
        "requirements": "Java, Spring Boot, Microservices, REST APIs, MySQL, AWS, Kafka",
        "score": 9,
        "status": "Not Applied",
        "notes": "Fintech unicorn; payment systems, high concurrency - strong profile match"
    },
    {
        "title": "Java Spring Boot Developer",
        "company": "Infosys",
        "location": "Hybrid - Pune, Maharashtra",
        "url": "https://www.naukri.com/spring-boot-jobs",
        "date": "May 2026",
        "salary": "₹10,00,000 - ₹18,00,000",
        "requirements": "Java, Spring Boot, REST APIs, AWS, Docker, MySQL, Agile, JUnit",
        "score": 8,
        "status": "Not Applied",
        "notes": "Global IT; stable career path, multiple project domains available"
    },
    {
        "title": "AWS Cloud Engineer (Java)",
        "company": "Amazon Web Services",
        "location": "Onsite - Hyderabad, Telangana",
        "url": "https://aws.amazon.com/careers/",
        "date": "May 2026",
        "salary": "₹30,00,000 - ₹52,00,000",
        "requirements": "Java, AWS (EC2, S3, DynamoDB, Lambda, IAM), Spring Boot, Microservices, CI/CD",
        "score": 10,
        "status": "Not Applied",
        "notes": "Dream match - AWS expertise + Java + prior Amazon internship is a huge advantage"
    },
    {
        "title": "Senior Java Developer (Remote)",
        "company": "Senior JAVA Spring Boot Backend (via Wellfound)",
        "location": "Remote - India",
        "url": "https://wellfound.com/role/l/spring-boot/india",
        "date": "May 2026",
        "salary": "₹15,00,000 - ₹30,00,000",
        "requirements": "Java, Spring Boot, REST APIs, AWS, Docker, MySQL, OAuth 2.0",
        "score": 8,
        "status": "Not Applied",
        "notes": "Fully remote startup role; OAuth 2.0 experience is a differentiator"
    },
    {
        "title": "Backend Engineer (Java/Spring Boot)",
        "company": "Freshworks",
        "location": "Hybrid - Chennai, Tamil Nadu",
        "url": "https://in.linkedin.com/jobs/spring-boot-java-developer-jobs",
        "date": "May 2026",
        "salary": "₹20,00,000 - ₹35,00,000",
        "requirements": "Java, Spring Boot, Microservices, AWS, Kafka, Docker, Kubernetes",
        "score": 8,
        "status": "Not Applied",
        "notes": "SaaS product company; good work-life balance, strong engineering culture"
    },
    {
        "title": "Software Engineer - Microservices",
        "company": "Citi",
        "location": "Hybrid - Pune, Maharashtra",
        "url": "https://www.glassdoor.co.in/Job/java-developer-with-aws-and-spring-boot-experience-jobs-SRCH_KO0,50.htm",
        "date": "May 2026",
        "salary": "₹18,00,000 - ₹32,00,000",
        "requirements": "Java, Spring Boot, Microservices, AWS, REST APIs, Docker, CI/CD, OAuth 2.0",
        "score": 8,
        "status": "Not Applied",
        "notes": "Global bank; enterprise-grade systems, security (OAuth/RBAC) skills valued"
    },
    {
        "title": "Java Backend Developer",
        "company": "Zerodha",
        "location": "Remote - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/aws-java-jobs",
        "date": "May 2026",
        "salary": "₹20,00,000 - ₹40,00,000",
        "requirements": "Java, Spring Boot, Microservices, REST APIs, MySQL, AWS, Docker",
        "score": 8,
        "status": "Not Applied",
        "notes": "Bootstrapped fintech; lean team, high ownership, strong remote culture"
    },
    {
        "title": "SDE (Java + AWS)",
        "company": "TriSys IT Solutions",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://offcampusjobs4u.com/trisys-recruitment-2026-java-developer-bengaluru-fresher/",
        "date": "May 2026",
        "salary": "₹6,00,000 - ₹10,00,000",
        "requirements": "Java, Spring Boot, REST APIs, SQL, AWS basics",
        "score": 6,
        "status": "Not Applied",
        "notes": "Entry-level posting; candidate is overqualified but could negotiate higher role"
    },
    {
        "title": "Cloud Software Engineer (Java)",
        "company": "Google Cloud (via LinkedIn)",
        "location": "Onsite - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/aws-developer-jobs",
        "date": "May 2026",
        "salary": "₹35,00,000 - ₹60,00,000",
        "requirements": "Java, Cloud (AWS/GCP), Microservices, Docker, Kubernetes, CI/CD, System Design",
        "score": 7,
        "status": "Not Applied",
        "notes": "Top-tier comp; AWS skills transferable to GCP, but GCP-specific knowledge preferred"
    },
    {
        "title": "Backend Engineer (Java Microservices)",
        "company": "Paytm",
        "location": "Hybrid - Noida, Uttar Pradesh",
        "url": "https://in.linkedin.com/jobs/java-developer-(spring-boot-and-microservices)-jobs",
        "date": "May 2026",
        "salary": "₹18,00,000 - ₹30,00,000",
        "requirements": "Java, Spring Boot, Kafka, Microservices, AWS, MySQL, Docker, REST APIs",
        "score": 9,
        "status": "Not Applied",
        "notes": "Fintech; high-volume transaction systems, Kafka + microservices experience valued"
    },
    {
        "title": "Software Engineer - Java (Backend)",
        "company": "Walmart Global Tech",
        "location": "Hybrid - Bengaluru, Karnataka",
        "url": "https://in.linkedin.com/jobs/software-engineer-java-spring-boot-kafka-jobs-bengaluru",
        "date": "May 2026",
        "salary": "₹22,00,000 - ₹40,00,000",
        "requirements": "Java, Spring Boot, Kafka, Microservices, AWS, Docker, CI/CD, System Design",
        "score": 9,
        "status": "Not Applied",
        "notes": "Enterprise scale; Spring Boot + Kafka combo is a direct match"
    },
    {
        "title": "Java Developer (Spring Boot + Microservices)",
        "company": "Capgemini",
        "location": "Hybrid - Gurugram, Haryana",
        "url": "https://in.linkedin.com/jobs/spring-boot-jobs-gurugram",
        "date": "May 2026",
        "salary": "₹10,00,000 - ₹20,00,000",
        "requirements": "Java, Spring Boot, Microservices, REST APIs, AWS, MySQL, Docker, Agile",
        "score": 8,
        "status": "Not Applied",
        "notes": "IT consulting; diverse project exposure, steady growth path"
    },
]

for row_idx, job in enumerate(jobs, 2):
    ws.cell(row=row_idx, column=1, value=job["title"]).border = thin_border
    ws.cell(row=row_idx, column=2, value=job["company"]).border = thin_border
    ws.cell(row=row_idx, column=3, value=job["location"]).border = thin_border

    url_cell = ws.cell(row=row_idx, column=4, value=job["url"])
    url_cell.hyperlink = job["url"]
    url_cell.font = Font(color="0563C1", underline="single")
    url_cell.border = thin_border

    ws.cell(row=row_idx, column=5, value=job["date"]).border = thin_border
    ws.cell(row=row_idx, column=6, value=job["salary"]).border = thin_border
    ws.cell(row=row_idx, column=7, value=job["requirements"]).border = thin_border

    score_cell = ws.cell(row=row_idx, column=8, value=job["score"])
    score_cell.alignment = Alignment(horizontal="center")
    score_cell.border = thin_border
    if job["score"] >= 9:
        score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif job["score"] >= 7:
        score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.cell(row=row_idx, column=9, value=job["status"]).border = thin_border
    ws.cell(row=row_idx, column=10, value=job["notes"]).border = thin_border

col_widths = [38, 35, 32, 55, 12, 28, 65, 16, 18, 65]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = "A2"

wb.save(filename)
print(f"Created: {filename}")
